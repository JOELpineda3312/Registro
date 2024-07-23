from flask import Flask, request, jsonify, send_from_directory
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side
import os
from datetime import datetime
from flask_cors import CORS

app = Flask(__name__)
CORS(app)

PORT = 8080

# Ruta para servir el archivo HTML
@app.route('/')
def index():
    return send_from_directory('templates', 'index.html')

# Leer empleados desde el archivo de Excel
def cargar_empleados():
    filePath = './BaseDatosRegistro.xlsx'
    empleados = []
    if os.path.exists(filePath):
        workbook = load_workbook(filePath)
        if 'Empleados' in workbook.sheetnames:
            worksheet = workbook['Empleados']
            for row in worksheet.iter_rows(min_row=2, max_col=4, values_only=True):
                empleados.append({
                    'Nombres': row[0],
                    'Tipo documento': row[1],
                    'No. Documento': row[2],
                    'Rol': row[3]
                })
    return empleados

@app.route('/validar/<cedula>', methods=['GET'])
def validar(cedula):
    try:
        empleados = cargar_empleados()
        empleado = next((emp for emp in empleados if str(emp['No. Documento']) == cedula), None)
        if empleado:
            return jsonify({'success': True, **empleado})
        else:
            return jsonify({'success': False, 'message': 'Empleado no encontrado'}), 404
    except Exception as e:
        return jsonify({'success': False, 'message': 'Error en la validación'}), 500

@app.route('/registrar', methods=['POST'])
def registrar():
    try:
        data = request.json
        numeroCedula = data.get('numeroCedula')
        tipoRegistro = data.get('tipoRegistro')

        empleados = cargar_empleados()
        empleado = next((emp for emp in empleados if str(emp['No. Documento']) == numeroCedula), None)
        if empleado:
            filePath = './BaseDatosRegistro.xlsx'
            workbook = load_workbook(filePath) if os.path.exists(filePath) else Workbook()

            if 'BaseDatos' in workbook.sheetnames:
                worksheet = workbook['BaseDatos']
            else:
                worksheet = workbook.active
                worksheet.title = 'BaseDatos'
                worksheet.append(['No. Registro', 'Nombres', 'Tipo Documento', 'No. Documento', 'Rol', 'Tipo Registro', 'Hora', 'Fecha'])

            num_registro = len(list(worksheet.iter_rows(min_row=4, max_col=1, values_only=True))) + 1
            newRow = [
                num_registro,
                empleado['Nombres'],
                empleado['Tipo documento'],
                empleado['No. Documento'],
                empleado['Rol'],
                tipoRegistro,
                datetime.now().strftime('%I:%M:%S %p'),
                datetime.now().strftime('%Y-%m-%d')
            ]

            tipo_color = {
                'LLEGADA': 'D0F0C0',  # Light Green
                'SALIDA': 'F3F3A8'    # Light Yellow
            }
            fill_color = tipo_color.get(tipoRegistro, 'FFFFFF')
            worksheet.append(newRow)

            row_num = len(list(worksheet.iter_rows(min_row=4, max_col=1, values_only=True))) + 3
            for cell in worksheet[row_num]:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(left=Side(style='thin'),
                                     right=Side(style='thin'),
                                     top=Side(style='thin'),
                                     bottom=Side(style='thin'))
                if cell.column == 6:  # Columna F
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')

            worksheet.row_dimensions[row_num].height = 30
            workbook.save(filePath)

            return jsonify({'success': True, 'message': f'Hola, Muchas Gracias por Registrar tu {tipoRegistro}. Ten un Feliz día.'})
        else:
            return jsonify({'success': False, 'message': 'Empleado no encontrado'}), 404
    except Exception as e:
        return jsonify({'success': False, 'message': 'Error en el registro'}), 500

if __name__ == '__main__':
    app.run(debug=True, port=PORT)
